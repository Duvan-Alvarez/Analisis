"""Utilidades para exportar análisis de CVs a Excel y PDF."""

import json
from io import BytesIO
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors


def export_to_excel(results: List[Dict], job_description: str = None) -> BytesIO:
    """Exporta análisis de CVs a un archivo Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis de CVs"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Encabezados
    headers = [
        "Ranking", "Nombre", "Email", "Teléfono", "Ubicación",
        "Score General", "Score Matching", "Años Experiencia",
        "Fortalezas", "Áreas Mejora"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Datos
    for row_idx, result in enumerate(results, 2):
        row_data = [
            result.get("rank", "N/A"),
            result.get("candidate_name", "N/A"),
            result.get("email", "N/A"),
            result.get("phone", "N/A"),
            result.get("location", "N/A"),
            result.get("overall_score", "N/A"),
            result.get("match_score", "N/A"),
            result.get("years_of_experience", "N/A"),
            "; ".join(result.get("strengths", [])),
            "; ".join(result.get("areas_for_improvement", []))
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 25
    ws.column_dimensions["J"].width = 25
    
    # Detalles por candidato en hojas adicionales
    for result in results:
        if result.get("candidate_name"):
            sheet = wb.create_sheet(title=result["candidate_name"][:31])
            row = 1
            
            # Información básica
            sheet[f"A{row}"] = "ANÁLISIS DE CV"
            sheet[f"A{row}"].font = Font(bold=True, size=14)
            row += 2
            
            sheet[f"A{row}"] = "Nombre:"
            sheet[f"B{row}"] = result.get("candidate_name", "N/A")
            row += 1
            
            sheet[f"A{row}"] = "Email:"
            sheet[f"B{row}"] = result.get("email", "N/A")
            row += 1
            
            sheet[f"A{row}"] = "Teléfono:"
            sheet[f"B{row}"] = result.get("phone", "N/A")
            row += 1
            
            sheet[f"A{row}"] = "Ubicación:"
            sheet[f"B{row}"] = result.get("location", "N/A")
            row += 2
            
            # Puntuaciones
            sheet[f"A{row}"] = "PUNTUACIONES"
            sheet[f"A{row}"].font = Font(bold=True, size=12)
            row += 1
            
            sheet[f"A{row}"] = "Score General:"
            sheet[f"B{row}"] = result.get("overall_score", "N/A")
            row += 1
            
            if result.get("match_score"):
                sheet[f"A{row}"] = "Score Matching:"
                sheet[f"B{row}"] = result.get("match_score", "N/A")
                row += 1
            
            # Habilidades
            row += 1
            sheet[f"A{row}"] = "HABILIDADES"
            sheet[f"A{row}"].font = Font(bold=True, size=12)
            row += 1
            
            skills = result.get("skills", {})
            if skills.get("technical"):
                sheet[f"A{row}"] = "Técnicas:"
                sheet[f"B{row}"] = ", ".join(skills["technical"])
                row += 1
            
            if skills.get("soft"):
                sheet[f"A{row}"] = "Blandas:"
                sheet[f"B{row}"] = ", ".join(skills["soft"])
                row += 1
            
            if skills.get("languages"):
                sheet[f"A{row}"] = "Idiomas:"
                sheet[f"B{row}"] = ", ".join(skills["languages"])
                row += 1
            
            # Educación
            row += 1
            sheet[f"A{row}"] = "EDUCACIÓN"
            sheet[f"A{row}"].font = Font(bold=True, size=12)
            row += 1
            
            for edu in result.get("education", []):
                sheet[f"A{row}"] = f"{edu.get('degree', 'N/A')} - {edu.get('institution', 'N/A')}"
                sheet[f"A{row}"].font = Font(italic=True)
                row += 1
            
            # Experiencia
            row += 1
            sheet[f"A{row}"] = "EXPERIENCIA"
            sheet[f"A{row}"].font = Font(bold=True, size=12)
            row += 1
            
            for exp in result.get("experience", []):
                sheet[f"A{row}"] = f"{exp.get('title', 'N/A')} @ {exp.get('company', 'N/A')}"
                sheet[f"A{row}"].font = Font(bold=True)
                row += 1
                sheet[f"A{row}"] = exp.get('duration', 'N/A')
                row += 1
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_pdf(results: List[Dict], job_description: str = None) -> BytesIO:
    """Exporta análisis de CVs a un archivo PDF."""
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#4472C4"),
        spaceAfter=30,
        alignment=1  # Centro
    )
    
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=colors.HexColor("#4472C4"),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Título
    story.append(Paragraph("Análisis de Candidatos - CVision", title_style))
    story.append(Spacer(1, 0.3 * inch))
    
    if job_description:
        story.append(Paragraph(f"<b>Oferta de trabajo:</b> {job_description[:100]}...", styles["Normal"]))
        story.append(Spacer(1, 0.2 * inch))
    
    # Tabla resumen
    data = [["Ranking", "Nombre", "Score", "Match", "Estado"]]
    for result in results:
        data.append([
            str(result.get("rank", "N/A")),
            result.get("candidate_name", "N/A")[:30],
            str(result.get("overall_score", "N/A")),
            str(result.get("match_score", "N/A")) if result.get("match_score") else "N/A",
            "✓" if result.get("overall_score", 0) > 70 else "○"
        ])
    
    table = Table(data, colWidths=[1*inch, 2*inch, 1*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    story.append(table)
    story.append(Spacer(1, 0.3 * inch))
    
    # Detalle de cada candidato
    for idx, result in enumerate(results, 1):
        if idx > 1:
            story.append(PageBreak())
        
        story.append(Paragraph(f"Candidato #{idx}: {result.get('candidate_name', 'N/A')}", heading_style))
        
        details = f"""
        <b>Email:</b> {result.get('email', 'N/A')}<br/>
        <b>Teléfono:</b> {result.get('phone', 'N/A')}<br/>
        <b>Ubicación:</b> {result.get('location', 'N/A')}<br/>
        <b>Score General:</b> {result.get('overall_score', 'N/A')}/100<br/>
        """
        
        if result.get("match_score"):
            details += f"<b>Score Matching:</b> {result.get('match_score', 'N/A')}/100<br/>"
        
        story.append(Paragraph(details, styles["Normal"]))
        story.append(Spacer(1, 0.1 * inch))
        
        # Fortalezas
        if result.get("strengths"):
            strengths_text = ", ".join(result["strengths"])
            story.append(Paragraph(f"<b>Fortalezas:</b> {strengths_text}", styles["Normal"]))
            story.append(Spacer(1, 0.1 * inch))
        
        # Habilidades
        if result.get("skills"):
            skills = result["skills"]
            if skills.get("technical"):
                story.append(Paragraph(f"<b>Habilidades Técnicas:</b> {', '.join(skills['technical'][:5])}", styles["Normal"]))
            if skills.get("soft"):
                story.append(Paragraph(f"<b>Habilidades Blandas:</b> {', '.join(skills['soft'][:3])}", styles["Normal"]))
            story.append(Spacer(1, 0.1 * inch))
    
    doc.build(story)
    output.seek(0)
    return output
